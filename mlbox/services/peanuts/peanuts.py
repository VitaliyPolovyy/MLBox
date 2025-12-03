import os
import base64
from io import BytesIO
from pathlib import Path
import traceback
from typing import List, Optional, Tuple
import cv2
import httpx
import numpy as np
import pandas as pd
from dotenv import load_dotenv
from huggingface_hub import hf_hub_download
from loguru import logger
from openpyxl.drawing.image import Image as OpenPyxlImage
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage
from mlbox.models.peanuts.cls.yolo_cls_model import YOLOPeanutsClassifier
from mlbox.models.peanuts.detection.yolo_detector_model import YOLOPeanutsDetector
from mlbox.models.peanuts.detection.unet_detector_model import UNetPeanutsDetector
from mlbox.services.peanuts.datatype import Ellipse, Status, PeanutProcessingRequest, PeanutProcessingResult, OnePeanutProcessingResult, BaseResponseJson, PeanutDataResponseJson
from mlbox.settings import ROOT_DIR, LOG_LEVEL
from mlbox.utils.cvtools import preprocess_images_with_white_rectangle
from mlbox.utils.logger import get_logger, get_artifact_service

CURRENT_DIR = Path(__file__).parent
SERVICE_NAME = "Peanuts"
app_logger = get_logger(ROOT_DIR)
artifact_service = get_artifact_service(ROOT_DIR)

# Load environment variables - first from service directory, then from global .env.mlbox
load_dotenv(CURRENT_DIR / ".env", override=False)  # Load local .env first if exists
env_file = Path.home() / "credentials" / ".env.mlbox"
if env_file.exists():
    load_dotenv(env_file, override=False)  # Then load global credentials (override=False to keep local values)

#os.environ["CURL_CA_BUNDLE"] = ""
HF_TOKEN = os.getenv("HF_TOKEN")
HF_PEANUT_SEG_REPO_ID = os.getenv("HF_PEANUT_SEG_REPO_ID")
HF_PEANUT_SEG_FILE = os.getenv("HF_PEANUT_SEG_FILE")
HF_PEANUT_SEG_SEPARATED_REPO_ID = os.getenv("HF_PEANUT_SEG_SEPARATED_REPO_ID")
HF_PEANUT_SEG_SEPARATED_FILE = os.getenv("HF_PEANUT_SEG_SEPARATED_FILE")
HF_PEANUT_CLS_REPO_ID = os.getenv("HF_PEANUT_CLS_REPO_ID")
HF_PEANUT_CLS_FILE = os.getenv("HF_PEANUT_CLS_FILE")
ERP_ENDPOINT = (
    "https://ite.roshen.com:4433/WS/api/_MLBOX_HANDLE_RESPONSE?call_in_async_mode=false"
)



input_folder = ROOT_DIR / "tmp" / CURRENT_DIR.name / "input"


def process_requests(
    requests: List[PeanutProcessingRequest],
) -> List[PeanutProcessingResult]:

    """
    Processes a list of peanut processing requests and returns the results.

    Args:
        requests (List[PeanutProcessingRequest]): A list of peanut processing requests.

    Returns:
        List[PeanutProcessingResult]: A list of results after processing the peanut requests.

    The function performs the following steps:
    1. Processes the peanut images from the requests.
    2. Formats the result based on client needs.
    3. Prepares an Excel file for each processed result.
    4. Updates the status and error message of each result.
    5. Posts the result to the client via a REST request.

    Each result includes the status, error message, and the filename of the generated Excel file.
    """

    peanut_processing_results = process_peanuts_images(requests)

    for request, peanut_processing_result in zip(requests, peanut_processing_results):

        # Step 2: Format result based on client needs
        excel_file = prepare_excel(peanut_processing_result)
        peanut_processing_result.status = Status.SUCCESS
        peanut_processing_result.error_message = "Successfully processed the request."
        peanut_processing_result.excel_filename = excel_file
        post_rest_request_to_client(
            ERP_ENDPOINT, excel_file, request.alias, request.key, request.image_filename
        )

    return peanut_processing_results


def process_peanuts_separated_segmentation(
    preprocessed_images: List[np.ndarray],
    sv_detections: List,
    separated_seg_model_path: str,
    padding: int = 10,
) -> List[List[np.ndarray]]:
    """
    Process each detected peanut separately using cropped segmentation.
    
    Args:
        preprocessed_images: List of preprocessed full images
        sv_detections: List of sv.Detections from full image (for bboxes)
        separated_seg_model_path: Path to separated segmentation model
        padding: Padding around bbox when cropping (default: 10)
        
    Returns:
        List of lists of masks (one list per image, one mask per peanut)
        Masks are in full image coordinates
    """
    app_logger.info(SERVICE_NAME, f"Loading separated segmentation model: {separated_seg_model_path}")
    separated_detector = YOLOPeanutsDetector(separated_seg_model_path)
    
    all_separated_masks = []
    
    for preprocessed_image, sv_detection in zip(preprocessed_images, sv_detections):
        separated_masks = []
        h, w = preprocessed_image.shape[:2]
        
        for index in range(len(sv_detection.xyxy)):
            xyxy = sv_detection.xyxy[index]
            x1, y1, x2, y2 = map(int, xyxy)
            
            # Add padding to crop
            x1_padded = max(0, x1 - padding)
            y1_padded = max(0, y1 - padding)
            x2_padded = min(w, x2 + padding)
            y2_padded = min(h, y2 + padding)
            
            # Crop peanut
            cropped_peanut = preprocessed_image[y1_padded:y2_padded, x1_padded:x2_padded]
            
            # Run segmentation on cropped image
            cropped_detections = separated_detector.detect(
                [cropped_peanut], verbose=False, imgsz=128, conf=0.5
            )
            
            if cropped_detections and len(cropped_detections) > 0 and len(cropped_detections[0].mask) > 0:
                # Get mask from cropped detection
                cropped_mask = cropped_detections[0].mask[0]
                cropped_h, cropped_w = cropped_mask.shape
                crop_h_actual, crop_w_actual = y2_padded - y1_padded, x2_padded - x1_padded
                
                # Resize mask if needed (in case model output size differs from crop size)
                if cropped_h != crop_h_actual or cropped_w != crop_w_actual:
                    cropped_mask = cv2.resize(
                        cropped_mask.astype(np.uint8),
                        (crop_w_actual, crop_h_actual),
                        interpolation=cv2.INTER_NEAREST
                    ).astype(bool)
                
                # Create full-size mask and translate coordinates
                full_mask = np.zeros((h, w), dtype=bool)
                full_mask[y1_padded:y2_padded, x1_padded:x2_padded] = cropped_mask
                separated_masks.append(full_mask)
            else:
                # Fallback: use bbox as mask if no detection
                app_logger.warning(SERVICE_NAME, f"No separated segmentation detected for peanut {index}, using bbox fallback")
                full_mask = np.zeros((h, w), dtype=bool)
                full_mask[y1:y2, x1:x2] = True
                separated_masks.append(full_mask)
        
        all_separated_masks.append(separated_masks)
    
    app_logger.info(SERVICE_NAME, f"Separated segmentation completed | images={len(all_separated_masks)}")
    return all_separated_masks


def process_peanuts_images(
    requests: List[PeanutProcessingRequest],
) -> List[PeanutProcessingResult]:

    app_logger.info(SERVICE_NAME, f"Processing peanuts images: {len(requests)}")
    peanut_processing_results: List[PeanutProcessingResult] = []

    input_images, input_images_filename = zip(
        *[(np.array(request.image), request.image_filename) for request in requests]
    )

    # prepare images for processing
    app_logger.debug(SERVICE_NAME, f"Preparing images for processing: {len(input_images)}")
    try:
        preprocessed_results = preprocess_images_with_white_rectangle(input_images= input_images, target_width = 2000)
        app_logger.debug(SERVICE_NAME, f"Preprocessed images: {len(preprocessed_results)}")
        preprocessed_images, pixels_per_mm_values = zip(*preprocessed_results)
        app_logger.debug(SERVICE_NAME, f"Pixels per mm values: {pixels_per_mm_values}")
    except ValueError as e:
        # If A4 paper detection fails, use simple resizing instead
        error_trace = traceback.format_exc()
        app_logger.error(SERVICE_NAME, f"A4 paper detection failed: {e}. Using simple resizing.\n{error_trace}")
        preprocessed_images = []
        pixels_per_mm_values = []
        for image in input_images:
            # Simple resize to target width while maintaining aspect ratio
            height, width = image.shape[:2]
            aspect_ratio = width / height
            new_height = int(2000 / aspect_ratio)
            resized_image = cv2.resize(image, (2000, new_height))
            preprocessed_images.append(resized_image)
            # Use a default pixels_per_mm value (this is approximate)
            pixels_per_mm_values.append(5.0)  # Default value, may need adjustment

    if app_logger.level == "DEBUG":
        save_images_with_annotations(
            preprocessed_images, step_name="preprocessing", output_folder=artifact_service.get_service_dir(SERVICE_NAME) )

    # load weights file and detect peanuts on the images
    app_logger.info(SERVICE_NAME, f"Starting classifier model download | repo_id={HF_PEANUT_CLS_REPO_ID} | file={HF_PEANUT_CLS_FILE}")
    try:
        cls_model_path = hf_hub_download(
            repo_id=HF_PEANUT_CLS_REPO_ID, filename=HF_PEANUT_CLS_FILE, token=HF_TOKEN
        )
        app_logger.info(SERVICE_NAME, f"Classifier model downloaded: {cls_model_path}")
    except Exception as e:
        error_trace = traceback.format_exc()
        app_logger.error(SERVICE_NAME, f"Failed to download classifier model | repo_id={HF_PEANUT_CLS_REPO_ID} | error={str(e)}\n{error_trace}")
        raise
    
    app_logger.info(SERVICE_NAME, f"Initializing classifier model: {cls_model_path}")
    try:
        classifier = YOLOPeanutsClassifier(cls_model_path)
        app_logger.info(SERVICE_NAME, f"Classifier model initialized successfully")
    except Exception as e:
        error_trace = traceback.format_exc()
        app_logger.error(SERVICE_NAME, f"Failed to initialize classifier model | path={cls_model_path} | error={str(e)}\n{error_trace}")
        raise

    # load weights file and detect peanuts on the images
    app_logger.info(SERVICE_NAME, f"Starting detector model download | repo_id={HF_PEANUT_SEG_REPO_ID} | file={HF_PEANUT_SEG_FILE}")
    try:
        det_model_path = hf_hub_download(
            repo_id=HF_PEANUT_SEG_REPO_ID, filename=HF_PEANUT_SEG_FILE, token=HF_TOKEN
        )
        app_logger.info(SERVICE_NAME, f"Detector model downloaded: {det_model_path}")
    except Exception as e:
        error_trace = traceback.format_exc()
        app_logger.error(SERVICE_NAME, f"Failed to download detector model | repo_id={HF_PEANUT_SEG_REPO_ID} | error={str(e)}\n{error_trace}")
        raise
    
    app_logger.info(SERVICE_NAME, f"Initializing detector model: {det_model_path}")
    try:
        detector = YOLOPeanutsDetector(det_model_path)
        app_logger.info(SERVICE_NAME, f"Detector model initialized successfully")
    except Exception as e:
        error_trace = traceback.format_exc()
        app_logger.error(SERVICE_NAME, f"Failed to initialize detector model | path={det_model_path} | error={str(e)}\n{error_trace}")
        raise
    
    # Load separated segmentation model (UNet) from Hugging Face, as before
    separated_detector = None
    if HF_PEANUT_SEG_SEPARATED_REPO_ID and HF_PEANUT_SEG_SEPARATED_FILE:
        
        try:
            separated_seg_model_path = hf_hub_download(
                repo_id=HF_PEANUT_SEG_SEPARATED_REPO_ID,
                filename=HF_PEANUT_SEG_SEPARATED_FILE,
                token=HF_TOKEN,
            )
            separated_detector = UNetPeanutsDetector(separated_seg_model_path)
        except Exception as e:
            error_trace = traceback.format_exc()
            app_logger.error(SERVICE_NAME, f"Failed to load separated segmentation model | error={str(e)}\n{error_trace}",)
            raise
    
    app_logger.info(SERVICE_NAME, f"Starting peanut detection | images_count={len(preprocessed_images)}")
    try:
        sv_detections = detector.detect(
            preprocessed_images, verbose=True, imgsz=1024, conf=0.6
        )
        app_logger.info(SERVICE_NAME, f"Detection completed | detected_objects={sum(len(det.xyxy) for det in sv_detections)}")

        # Apply NMS to remove overlapping boxes
        nms_iou_threshold = 0.5
        sv_detections = [det.with_nms(threshold=nms_iou_threshold) for det in sv_detections]
        app_logger.info(SERVICE_NAME, f"NMS applied | remaining_objects={sum(len(det.xyxy) for det in sv_detections)}")

    except Exception as e:
        error_trace = traceback.format_exc()
        app_logger.error(SERVICE_NAME, f"Detection failed | error={str(e)}\n{error_trace}")
        raise

    if LOG_LEVEL == "DEBUG":
        bounding_boxes = [detection.xyxy for detection in sv_detections]
        masks = [detection.mask for detection in sv_detections]
        save_images_with_annotations(
            preprocessed_images,
            step_name="detection",
            output_folder=artifact_service.get_service_dir(SERVICE_NAME),
            bounding_boxes=bounding_boxes,
            masks=masks,
        )

    # for each image and each peanut: crop peanut, fit elipse, classify, fill result
    app_logger.info(SERVICE_NAME, f"Processing {len(preprocessed_images)} images through classification pipeline")
    for img_idx, (input_image_filename, preprocessed_image, sv_detection, pixels_per_mm) in enumerate(
        zip(input_images_filename, preprocessed_images, sv_detections, pixels_per_mm_values), 1
    ):
        app_logger.info(SERVICE_NAME, f"Processing image {img_idx}/{len(preprocessed_images)}: {input_image_filename} | detected_peanuts={len(sv_detection.xyxy)}")

        peanuts: List[OnePeanutProcessingResult] = []
        one_peanut_images: List[np.ndarray] = []

        sorted_indices = sorted(
            range(len(sv_detection.xyxy)),
            key=lambda idx: (sv_detection.xyxy[idx][1], sv_detection.xyxy[idx][0]),
        )
        ordered_index = 0
        for index in sorted_indices:

            xyxy = sv_detection.xyxy[index]
            mask = sv_detection.mask[index]
            x1, y1, x2, y2 = map(int, xyxy)
            # Crop peanut image
            one_peanut_image = preprocessed_image[y1:y2, x1:x2].copy()

            # Set outside pixels to white
            one_peanut_image[mask[y1:y2, x1:x2] is False] = 255

            # Save the cropped peanut image
            if LOG_LEVEL == "DEBUG":
                pil_image = PILImage.fromarray(one_peanut_image)
                original_stem = Path(input_image_filename).stem
                artifact_service.save_artifact(
                    service=SERVICE_NAME,
                    # Use ordered_index so filenames, Excel rows and comparison labels match
                    file_name=f"{original_stem}_{ordered_index}.jpg",
                    data=pil_image,
                    sub_folder="temp"
                )

            one_peanut_images.append(one_peanut_image)

            # find contour and fit elipse
            contours, _ = cv2.findContours(
                mask.astype(np.uint8), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_NONE
            )
            contour = max(contours, key=cv2.contourArea)
            contour_reshaped = contour.reshape(-1, 2)
            center, axes, angle = cv2.fitEllipse(contour_reshaped)
            axes = (axes[0] * 0.9, axes[1] * 0.9)
            elipse = Ellipse(center=center, axes=axes, angle=angle)

            mask_separated = None
            contour_separated = None
            ellipse_separated = None
            
            # UNet expects RGB images (training & assessment use RGB),
            # while preprocessed_image/one_peanut_image are BGR (OpenCV).
            cropped_detections = separated_detector.detect(
                [one_peanut_image], verbose=False, imgsz=128, conf=0.5
            )
            
            if (cropped_detections and 
                len(cropped_detections) > 0 and 
                cropped_detections[0].mask is not None and 
                len(cropped_detections[0].mask) > 0):
                # Get mask from cropped detection
                cropped_mask = cropped_detections[0].mask[0]
                cropped_h, cropped_w = cropped_mask.shape
                crop_h_actual, crop_w_actual = y2 - y1, x2 - x1
                
                # Resize mask if needed (model output might differ from crop size)
                if cropped_h != crop_h_actual or cropped_w != crop_w_actual:
                    cropped_mask = cv2.resize(
                        cropped_mask.astype(np.uint8),
                        (crop_w_actual, crop_h_actual),
                        interpolation=cv2.INTER_NEAREST
                    ).astype(bool)
                
                # Create full-size mask and translate coordinates
                h, w = preprocessed_image.shape[:2]
                full_mask = np.zeros((h, w), dtype=bool)
                full_mask[y1:y2, x1:x2] = cropped_mask
                mask_separated = full_mask
                
                # Find contour from full mask (already in full image coordinates)
                contours_separated, _ = cv2.findContours(
                    mask_separated.astype(np.uint8), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_NONE
                )
                if contours_separated:
                    contour_separated = max(contours_separated, key=cv2.contourArea)
                    contour_separated_reshaped = contour_separated.reshape(-1, 2)
                    center_separated, axes_separated, angle_separated = cv2.fitEllipse(contour_separated_reshaped)
                    axes_separated = (axes_separated[0] * 0.9, axes_separated[1] * 0.9)
                    ellipse_separated = Ellipse(center=center_separated, axes=axes_separated, angle=angle_separated)

            peanut = OnePeanutProcessingResult(
                index=ordered_index,
                xyxy=xyxy,
                mask=mask,
                mask_separated=mask_separated,
                contour=contour,
                contour_separated=contour_separated,
                det_confidence=sv_detection.confidence[index],
                image=PILImage.fromarray(one_peanut_image),
                ellipse=elipse,
                ellipse_separated=ellipse_separated,
            )
            ordered_index = ordered_index + 1

            peanuts.append(peanut)

        app_logger.info(SERVICE_NAME, f"Classifying {len(one_peanut_images)} detected peanuts for image: {input_image_filename}")
        try:
            sv_cls = classifier.classify(one_peanut_images, verbose=False)
            app_logger.info(SERVICE_NAME, f"Classification completed for image: {input_image_filename}")
        except Exception as e:
            error_trace = traceback.format_exc()
            app_logger.error(SERVICE_NAME, f"Classification failed for image: {input_image_filename} | error={str(e)}\n{error_trace}")
            raise

        for index, classification in enumerate(sv_cls):
            peanuts[index].class_id = classification.class_id
            peanuts[index].class_confidence = classification.confidence

        peanut_processing_results.append(
            PeanutProcessingResult(
                peanuts=peanuts,
                weight_g=100,
                pixels_per_mm=pixels_per_mm,
                original_image=PILImage.fromarray(preprocessed_image),
                original_image_filename=input_image_filename,
                status=Status.SUCCESS,
                message="Successfully processed the request",
            )
        )

        # Ensure result folder exists
        
        for result in peanut_processing_results:
            result.result_image.save(
                artifact_service.get_service_dir(SERVICE_NAME) / f"result_{result.original_image_filename}",
                format="JPEG",
            )

    if LOG_LEVEL == "DEBUG":
        classes: List[List[str]] = []
        bounding_boxes: List[List[Tuple[int, int, int, int]]] = []
        masks: List[List[np.ndarray]] = []
        ellipses: List[List[Ellipse]] = []
        contours: List[List[np.ndarray]] = []
        indexes: List[List[int]] = []

        for peanut_result in peanut_processing_results:
            bounding_boxes.append([peanut.xyxy for peanut in peanut_result.peanuts])
            masks.append([peanut.mask for peanut in peanut_result.peanuts])
            classes.append([peanut.class_id for peanut in peanut_result.peanuts])
            ellipses.append([peanut.ellipse for peanut in peanut_result.peanuts])
            contours.append([peanut.contour for peanut in peanut_result.peanuts])
            indexes.append([peanut.index for peanut in peanut_result.peanuts])

        save_images_with_annotations(
            images=preprocessed_images,
            step_name="result",
            indexes=indexes,
            bounding_boxes=bounding_boxes,
            masks=masks,
            classes=classes,
            contours=contours,
            ellipses=ellipses,
            output_folder=artifact_service.get_service_dir(SERVICE_NAME),
        )

    return peanut_processing_results

def prepare_excel(peanut_processing_result: PeanutProcessingResult) -> Path:
    """
    Create an Excel file for the given PeanutProcessingResult.

    Args:
        peanut_processing_result (PeanutProcessingResult): The result to prepare the Excel file for.

    Returns:
        Path: The path to the generated Excel file.
    """
    
    # Define output file path
    excel_file = (
        artifact_service.get_service_dir(SERVICE_NAME)
        / f"{Path(peanut_processing_result.original_image_filename).stem}.xlsx"
    )

    if excel_file.exists():
        excel_file.unlink()

    # Create Excel writer
    with pd.ExcelWriter(excel_file, engine="openpyxl") as writer:
        # Prepare result table
        result_table = pd.DataFrame(
            [
                {
                    "№ п/п": idx,
                    "max діаметр, мм": (
                        round(
                            peanut.ellipse.axes[1]
                            / peanut_processing_result.pixels_per_mm,
                            1,
                        )
                        if peanut.ellipse
                        else None
                    ),
                    "min діаметр, мм": (
                        round(
                            peanut.ellipse.axes[0]
                            / peanut_processing_result.pixels_per_mm,
                            1,
                        )
                        if peanut.ellipse
                        else None
                    ),
                    "клас": YOLOPeanutsClassifier.class_names[peanut.real_class[0]],
                    "впевненність (клас)": round(peanut.real_class[1], 2),
                    "впевненність (маска)": round(peanut.det_confidence, 2),
                }
                for idx, peanut in enumerate(peanut_processing_result.peanuts)
            ]
        )

        # Write the result table to the first sheet
        result_table.to_excel(writer, sheet_name="Peanut Analysis", index=False)

        # Create a DataFrame for the calculated indicators
        indicators_table = pd.DataFrame(
            {
                "Показник": [
                    "Вага зразку, г",
                    "середньо кв.відх. (по 'меньша осі')",
                    "коефіціент варіації (по 'меньша осі'), %",
                    "середньо кв.відх. (по 'більша ось / меньша ось')",
                    "коефіціент варіації (по 'більша ось / меньша ось'), %",
                    "Шт в 1 унції",
                ],
                "Значення": [
                    peanut_processing_result.weight_g,
                    peanut_processing_result.standard_deviation_minor_axe,
                    peanut_processing_result.coefficient_variation_minor_axe,
                    peanut_processing_result.standard_deviation_ratio_axes,
                    peanut_processing_result.coefficient_variation_ratio_axes,
                    round(len(peanut_processing_result.peanuts) * 28.35 / 100, 2),
                ],
            }
        )

        indicators_table_start_col = result_table.shape[1] + 1
        # Write the indicators table to the second sheet starting from column E (5th column)
        indicators_table.to_excel(
            writer,
            sheet_name="Peanut Analysis",
            index=False,
            startcol=indicators_table_start_col,
        )

        # Access the workbook and worksheet
        worksheet = writer.sheets["Peanut Analysis"]

        # Set the alignment for all cells to wrap text

        worksheet.column_dimensions[
            get_column_letter(indicators_table_start_col + 1)
        ].width = 50

        for row in worksheet.iter_rows():
            for cell in row:

                if cell.column == indicators_table_start_col + 1:
                    cell.alignment = Alignment(wrap_text=True)
                else:
                    cell.alignment = Alignment(
                        wrap_text=True, horizontal="center", vertical="center"
                    )

        # Insert the result image into the second sheet
        result_image_sheet = writer.book.create_sheet(title="Result Image")
        image_stream = BytesIO()
        peanut_processing_result.result_image.save(image_stream, format="PNG")
        image_stream.seek(0)
        openpyxl_image = OpenPyxlImage(image_stream)
        openpyxl_image.width = openpyxl_image.width // 2
        openpyxl_image.height = openpyxl_image.height // 2
        result_image_sheet.add_image(openpyxl_image, "A1")

    # Add comparison sheet if separated segmentation is available
    if any(peanut.mask_separated is not None for peanut in peanut_processing_result.peanuts):
        # Convert PIL Image (RGB) to numpy array and then to BGR for OpenCV
        original_array = np.array(peanut_processing_result.original_image)
        if len(original_array.shape) == 3 and original_array.shape[2] == 3:
            # PIL Image is RGB, convert to BGR for OpenCV
            original_array = cv2.cvtColor(original_array, cv2.COLOR_RGB2BGR)
        comparison_image = create_comparison_image(
            original_array,
            peanut_processing_result.peanuts
        )
        
        # Save comparison image to temp folder
        comparison_filename = f"comparison_{Path(peanut_processing_result.original_image_filename).stem}.jpg"
        artifact_service.save_artifact(
            service=SERVICE_NAME,
            file_name=comparison_filename,
            data=comparison_image
        )
        
        comparison_sheet = writer.book.create_sheet(title="Comparison")
        image_stream = BytesIO()
        comparison_image.save(image_stream, format="PNG")
        image_stream.seek(0)
        openpyxl_image = OpenPyxlImage(image_stream)
        openpyxl_image.width = openpyxl_image.width // 2
        openpyxl_image.height = openpyxl_image.height // 2
        comparison_sheet.add_image(openpyxl_image, "A1")

    return excel_file


def create_comparison_image(
    preprocessed_image: np.ndarray,
    peanuts: List[OnePeanutProcessingResult],
) -> PILImage.Image:
    """
    Create comparison image with transparent masks for both segmentation approaches.
    
    Args:
        preprocessed_image: Original preprocessed image (BGR format from OpenCV)
        peanuts: List of peanut processing results
        
    Returns:
        PIL Image with both masks overlaid (red for full image, green for separated)
    """
    # Start with original image (BGR format)
    result_image = preprocessed_image.copy()
    
    # Ensure image is BGR (3 channels)
    if len(result_image.shape) == 2:
        result_image = cv2.cvtColor(result_image, cv2.COLOR_GRAY2BGR)
    elif result_image.shape[2] == 4:
        result_image = cv2.cvtColor(result_image, cv2.COLOR_BGRA2BGR)
    
    h, w = result_image.shape[:2]
    
    # Draw shapes and indexes for each peanut
    for peanut in peanuts:
        # Draw OLD ellipse (from original mask) in red
        if peanut.ellipse is not None:
            center = (int(peanut.ellipse.center[0]), int(peanut.ellipse.center[1]))
            axes = (
                int(peanut.ellipse.axes[0] / 2),
                int(peanut.ellipse.axes[1] / 2),
            )
            angle = peanut.ellipse.angle
            cv2.ellipse(
                result_image,
                center,
                axes,
                angle,
                0,
                360,
                (0, 0, 255),  # red ellipse for old mask
                2,
            )

        # Draw separated segmentation contour (green) – NEW mask
        if peanut.contour_separated is not None:
            cv2.drawContours(
                result_image,
                [peanut.contour_separated],
                -1,
                (0, 255, 0),  # green contour for new mask
                2,
            )

            # Draw rotated bounding box (yellow) from separated contour
            rect = cv2.minAreaRect(peanut.contour_separated)
            box = cv2.boxPoints(rect).astype(int)
            cv2.drawContours(
                result_image,
                [box],
                0,
                (0, 255, 255),  # yellow rotated bbox on new mask
                2,
            )

        # Draw peanut index near its original (YOLO) contour as anchor
        if peanut.contour is not None:
            moments = cv2.moments(peanut.contour)
            if moments["m00"] != 0:
                cx = int(moments["m10"] / moments["m00"])
                cy = int(moments["m01"] / moments["m00"])
                cv2.putText(
                    result_image,
                    str(peanut.index),
                    (cx, cy),
                    cv2.FONT_HERSHEY_SIMPLEX,
                    0.6,
                    (255, 255, 255),
                    2,
                    cv2.LINE_AA,
                )
    
    # Convert BGR to RGB for PIL
    result_image = cv2.cvtColor(result_image, cv2.COLOR_BGR2RGB)
    
    return PILImage.fromarray(result_image)


def post_rest_request_to_client(
    endpoint: str, excel_file: Path, alias: str, key: str, image_filename: str
) -> None:
    """
    Send the generated Excel file to the ERP endpoint.

    Args:
        excel_file (Path): Path to the Excel file to be sent.
        communication_data (Dict[str, str]): Communication details, including alias and alias_key.
    """
    headers = {"accept": "application/xml", "Content-Type": "application/json"}

    peanut_data = PeanutDataResponseJson(
        alias=alias,
        key=key,
        image_filename=image_filename,
        excel_file=base64.b64encode(excel_file.read_bytes()).decode("utf-8"),
    )

    peanut_response = BaseResponseJson(
        status="Success",
        message="Successfully sent the result to ERP.",
        service_name=SERVICE_NAME,
        timestamp=pd.Timestamp.now().isoformat(),
        data=peanut_data.to_json()  # pylint: disable=no-member
    )
    
    # Send the request
    response = httpx.post(endpoint, headers=headers, json={ "responseJson": peanut_response.to_json() }) # pylint: disable=no-member

    # Check response status
    if response.status_code == 200:
        app_logger.info(SERVICE_NAME, "Successfully sent the result to ERP.")
    else:
        app_logger.error(SERVICE_NAME, f"Failed to send the result to ERP. Status code: {response.status_code}, Response: {response.text}")
        raise Exception(f"Failed to send the result to ERP. Status code: {response.status_code}, Response: {response.text}")
    

def save_images_with_annotations(
    images: List[np.ndarray],
    step_name: str,
    output_folder: Path,
    bounding_boxes: Optional[List[List[Tuple[int, int, int, int]]]] = None,
    masks: Optional[List[List[np.ndarray]]] = None,
    contours: Optional[List[List[np.ndarray]]] = None,
    ellipses: Optional[List[List[Ellipse]]] = None,
    indexes: Optional[List[List[int]]] = None,
    classes: Optional[List[List[str]]] = None,
) -> None:

    output_folder.mkdir(parents=True, exist_ok=True)
    for i, img in enumerate(images):
        cv_image = img.copy()

        if bounding_boxes is not None and len(bounding_boxes) > i and False:
            for j, box in enumerate(bounding_boxes[i]):
                cv2.rectangle(
                    cv_image, (box[0], box[1]), (box[2], box[3]), (0, 255, 0), 2
                )
                label = (
                    f"{indexes[i][j]}"
                    if indexes is not None and len(indexes) > i
                    else ""
                )
                if (
                    classes is not None
                    and len(classes) > i
                    and classes[i][j] is not None
                ):
                    label += f" {classes[i][j]}"
                cv2.putText(
                    cv_image,
                    label,
                    (box[0], box[1] - 10),
                    cv2.FONT_HERSHEY_SIMPLEX,
                    0.5,
                    (0, 0, 255),
                    2,
                )

        if masks is not None and len(masks) > i and False:
            for mask in masks[i]:
                # Create an RGBA image with the mask and set the color to green with 50% transparency
                pil_mask = PILImage.fromarray(mask).convert("L")
                green_mask = PILImage.new(
                    "RGBA", pil_mask.size, (245, 0, 70, 255)
                )  # Green with 50% transparency
                pil_mask_rgba = PILImage.composite(
                    green_mask, PILImage.new("RGBA", pil_mask.size), pil_mask
                )

                # Convert RGBA mask to BGR format for OpenCV
                mask_bgr = np.array(pil_mask_rgba.convert("RGB"))
                mask_bgr = cv2.cvtColor(mask_bgr, cv2.COLOR_RGB2BGR)

                # Overlay the mask on the original image
                cv_image = cv2.addWeighted(cv_image, 1, mask_bgr, 0.9, 0)

        if ellipses is not None and len(ellipses) > i:
            for ellipse in ellipses[i]:
                center = (int(ellipse.center[0]), int(ellipse.center[1]))
                axes = (int(ellipse.axes[0] / 2), int(ellipse.axes[1] / 2))
                angle = ellipse.angle
                cv2.ellipse(
                    cv_image, center, axes, angle, 0, 360, (255, 0, 0), 2
                )  # Blue color with thickness 2

        if contours is not None and len(contours) > i and False and False:
            for contour in contours[i]:
                cv2.drawContours(
                    cv_image, [contour], -1, (0, 0, 255), 2
                )  # Hard blue color with thickness 2

        # Convert NumPy array back to PIL image
        pil_image = PILImage.fromarray(cv_image)

        #pil_image.save(file_name, format="JPEG")
        #artifact_service.save_artifact(service=SERVICE_NAME,file_name=f"{step_name}_{i}.jpg",data=pil_image)


def preprocessing_images_for_dataset(input_folder: Path, output_folder: Path) -> None:
    """
    Preprocess images from the input folder and save the preprocessed images to the output folder.

    Args:
        input_folder (Path): The folder containing the raw input images.
        output_folder (Path): The folder to save the preprocessed images.
    """
    # Gather all image files from the input folder
    output_files = {file.name for file in output_folder.glob("*.*")}
    image_files = [file for file in input_folder.glob("*.*") if file.name not in output_files]

    # Ensure the output folder exists
    output_folder.mkdir(parents=True, exist_ok=True)

    # Process each image one at a time
    for img_file in image_files:
        # Read the image and convert to numpy array
        input_image = np.array(PILImage.open(img_file))

        # Preprocess the image
        try:
            preprocessed_result = preprocess_images_with_white_rectangle(input_images=[input_image], target_width=2000)
        except ValueError as e:
            print(f"{e}")
            continue

        preprocessed_image = preprocessed_result[0][0]

        # Save the preprocessed image to the output folder
        output_path = output_folder / img_file.name
        PILImage.fromarray(preprocessed_image).save(output_path, format="JPEG")



def test_process_requests(input_folder: Path, output_folder: Path):
    """
    Test the process_requests function by gathering image files from the input folder,
    preparing PeanutProcessingRequest objects, and calling the main processing function.
    """

    # Gather all image files from the input folder
    image_files = list(
        input_folder.glob("31_10_2025_09_30.jpg")
    )  # You can filter by extension, e.g. "*.jpg" if needed

    # Prepare requests by reading each image
    requests = []
    for img_file in image_files:
        img = PILImage.open(img_file)
        requests.append(
            PeanutProcessingRequest(
                image=img,
                image_filename=img_file.name,
                alias="DMS",
                key="   9127673     1",
                response_method="HTTP_POST_REQUEST",
                response_endpoint="https://ite.roshen.com:4433/WS/api/_MLBOX_HANDLE_RESPONSE?call_in_async_mode=false",
            )
        )

    # Call the main processing function with the prepared requests
    process_requests(requests)




if __name__ == "__main__":
    
    #output_folder = Path(r"/mnt/c/My storage/Python projects/MLBox/assets/peanuts/datasets/preprocessed")
    #input_folder = Path(r"/mnt/c/My storage/Python projects/MLBox/assets/peanuts/datasets/Experiments-photo-lab/2025-02-19-phones")
    #preprocessing_images_for_dataset(input_folder, output_folder)




    input_folder = Path(r"/home/polovyi/projects/mlbox/assets/peanuts/datasets/2025-10-30-vkf-anton")
    output_folder = input_folder / "output"
    
    
    test_process_requests(input_folder, output_folder)
